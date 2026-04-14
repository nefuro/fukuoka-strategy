#!/usr/bin/env python3
"""
選挙管理委員会 Excel → CSV 変換ツール

prefectures/<県名>/raw/ にあるExcelファイルを読み込み、
prefectures/<県名>/data/ にCSVを出力します。

使い方:
  python3 tools/convert_excel.py prefectures/kanagawa

raw/ ディレクトリの命名規則（ファイル名にこれらのキーワードを含める）:
  - sangi + senkyoku                → 参院選挙区（候補者別開票区別）→ candidate.csv
  - sangi + hirei + district        → 参院比例（党派別開票区別）   → senate.csv
  - shugi + hirei + district        → 衆院比例（党派別開票区別）   → house.csv ※選管公開なしの可能性あり

注意:
  - 衆院比例の「党派別開票区別」データは多くの自治体で公開されておらず、
    PDFのみの場合があります。その場合は手作業で house.csv を作成してください。

必要パッケージ:
  pip install openpyxl
"""

import argparse
import csv
import glob
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl が必要です。 pip install openpyxl でインストールしてください。", file=sys.stderr)
    sys.exit(1)

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False


# ============================================================
# 共通ヘルパー
# ============================================================

# 政令指定都市（合計行はスキップし、区を子として採用）
DESIGNATED_CITIES = {
    '札幌市', '仙台市', 'さいたま市', '千葉市', '川崎市', '横浜市', '相模原市',
    '新潟市', '静岡市', '浜松市', '名古屋市', '京都市', '大阪市', '堺市',
    '神戸市', '岡山市', '広島市', '北九州市', '福岡市', '熊本市',
}

# 郡名パターン（神奈川県以外も含む一般的なパターン）
GUN_RE = re.compile(r'^[^\s]+郡')


def is_total_row(name):
    """合計行かどうかを判定"""
    s = str(name).strip()
    return s.endswith('計') or s in ('合計', '指定都市計', 'その他の市計', '町村計', '県計')


def is_gun(name):
    """郡名のみの行か（〜郡で終わる）"""
    s = str(name).strip()
    return bool(GUN_RE.match(s)) and not any(c in s for c in '町村市区')


def strip_gun_prefix(name):
    """町村名から郡名を除去する（例: 「足柄下郡箱根町」→「箱根町」）"""
    m = re.match(r'^(.+郡)(.+[町村])$', name)
    if m:
        return m.group(2)
    return name


def parse_number(val):
    """セル値を数値に変換する。文字列の場合はカンマ・空白を除去して変換。"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        try:
            return float(val.strip().replace(',', '').replace(' ', '').replace('\u3000', ''))
        except ValueError:
            return None
    return None


def find_col_by_keyword(ws, keyword, rows=(5, 7, 11)):
    """指定された行から、キーワードを含む列インデックスを探す"""
    for row in rows:
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val and keyword in str(val):
                return col
    return None


def find_total_col(ws, rows=(5, 7, 8, 11)):
    """合計列を探す"""
    for row in rows:
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                s = str(val).strip()
                if '得票数計' in s or '合計' in s or s == '計':
                    return col
    return None


# ============================================================
# 参院選挙区（候補者別開票区別）パーサー — g.xlsx形式
# ============================================================

def extract_area_name(col_a, col_b):
    """
    col_a と col_b から市区町村名を抽出する。
    返り値: 採用すべき地域名（政令指定都市の親行や郡の親行・合計行はNone）
    """
    a = str(col_a).strip() if col_a else ''
    b = str(col_b).strip() if col_b else ''

    # 合計行はスキップ
    if a and is_total_row(a):
        return None
    if b and is_total_row(b):
        return None

    # 政令指定都市の親行はスキップ（区が子として後続）
    if a and a in DESIGNATED_CITIES:
        return None

    # 郡名のみの親行はスキップ（町村が子として後続）
    if a and is_gun(a):
        return None

    # col_b に値がある場合（政令指定都市の区 or 郡の町村）
    if b:
        return strip_gun_prefix(b)

    # col_a に値があり、政令指定都市・郡・合計でない → 一般市
    if a:
        return strip_gun_prefix(a)

    return None


def parse_senate_district(filepath):
    """
    参院選挙区の候補者別開票区別Excel (g.xlsx形式) をパースする。
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    team_col = find_col_by_keyword(ws, 'チームみらい', rows=(5, 7))
    if team_col is None:
        return None
    total_col = find_total_col(ws, rows=(5, 7, 8))

    results = []
    for row in range(9, ws.max_row + 1):
        area = extract_area_name(
            ws.cell(row=row, column=1).value,
            ws.cell(row=row, column=2).value,
        )
        if not area:
            continue

        team_votes = parse_number(ws.cell(row=row, column=team_col).value)
        if team_votes is None:
            continue
        team_votes = round(team_votes)

        total = parse_number(ws.cell(row=row, column=total_col).value) if total_col else None
        rate = round(team_votes / total * 100, 1) if total and total > 0 else 0

        results.append({
            '地域': area,
            '得票': team_votes,
            '率': rate,
        })

    return results


# ============================================================
# 参院比例（党派別開票区別）パーサー — m.xlsx形式
# ============================================================

def parse_senate_proportional(filepath):
    """
    参院比例の党派別開票区別Excel (m.xlsx形式) をパースする。

    想定フォーマット:
    - Sheet: '開票区別'
    - Row 7: 党派名（各党3列ずつ: 得票総数, 政党等の得票総数, 名簿登載者を除く得票総数）
    - Row 12+: データ行
      - Col A: 親市区名（政令指定都市の合計行）
      - Col B: 市区町村名（フルネーム）
      - 各党の最初の列: 得票総数
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['開票区別'] if '開票区別' in wb.sheetnames else wb.active

    # チームみらいの列を探す（row 7）
    team_col = find_col_by_keyword(ws, 'チームみらい', rows=(7,))
    if team_col is None:
        return None

    # 全政党の「得票総数」列を集めて自治体別合計を算出する
    # row 7 で値があるセルが「政党の先頭列」
    party_cols = []
    for col in range(3, ws.max_column + 1):
        val = ws.cell(row=7, column=col).value
        if val and str(val).strip() and str(val).strip() not in ('党派名',):
            party_cols.append(col)

    results = []
    for row in range(12, ws.max_row + 1):
        area = extract_area_name(
            ws.cell(row=row, column=1).value,
            ws.cell(row=row, column=2).value,
        )
        if not area:
            continue

        team_votes = parse_number(ws.cell(row=row, column=team_col).value)
        if team_votes is None:
            continue
        team_votes = round(team_votes)

        # 全政党の得票総数を合計
        total = 0
        for pc in party_cols:
            v = parse_number(ws.cell(row=row, column=pc).value)
            if v is not None:
                total += v

        rate = round(team_votes / total * 100, 1) if total > 0 else 0

        results.append({
            '地域': area,
            '得票数': team_votes,
            '得票率': rate,
        })

    return results


# ============================================================
# 衆院比例（開票調シート）パーサー — r6hirei.xlsx形式
# ============================================================

def normalize_name(s):
    """全角空白・半角空白・no-break-space を全て除去"""
    if not s:
        return ''
    return str(s).replace(' ', '').replace('\u3000', '').replace('\xa0', '').strip()


_TOTAL_ROW_NAMES = {'県計', '指定市計', '一般市計', '市部計', '郡部計', '合計', '開票区名', '届出番号', '党派名'}


def extract_house_area(raw_name, parent_state):
    """
    衆院比例「開票調」シート用の地域名抽出。
    parent_state は [parent_city] という1要素リスト（状態保持）。
    """
    name = normalize_name(raw_name)
    if not name or name in _TOTAL_ROW_NAMES:
        return None

    # 政令指定都市の親行
    if name in DESIGNATED_CITIES:
        parent_state[0] = name
        return None

    # 区のみ（政令指定都市の区）
    if name.endswith('区') and parent_state[0]:
        return parent_state[0] + name

    # 「三浦郡葉山町」のように1セルに郡＋町
    m = re.match(r'^([^\s]+郡)(.+[町村])$', name)
    if m:
        return m.group(2)

    # 郡のみ
    if name.endswith('郡'):
        return None

    # 町村単独
    if name.endswith('町') or name.endswith('村'):
        return name

    # 一般市（区を含まない市）
    if name.endswith('市'):
        parent_state[0] = None  # 一般市が来たら親市リセット
        return name

    return None


def parse_house_proportional(filepath):
    """
    衆院比例の党派別開票区別データを「開票調」シートから抽出する。
    R6 (令和6年) のExcel形式に対応。R8 が同じフォーマットなら自動的に動作する。

    想定フォーマット:
    - Sheet: '開票調'
    - Row 6: 政党名（col 5〜）
    - Row 8 以降: 開票区別データ（2行ペア: 得票数 + 得票率）
      - col 1: 開票率
      - col 2: 開票区名
      - col 5〜14: 各党の得票数
      - col 15: 合計
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if '開票調' not in wb.sheetnames:
        return None
    ws = wb['開票調']

    # チームみらい列を探す（row 6）
    team_col = None
    for col in range(5, ws.max_column + 1):
        v = ws.cell(row=6, column=col).value
        if v and 'チームみらい' in str(v):
            team_col = col
            break
    if team_col is None:
        # R6 など過去選挙にはチームみらいがない → 警告
        print(f"  Warning: チームみらい列が見つかりません: {filepath}", file=sys.stderr)
        print(f"           （この選挙にチームみらい候補がいないか、R8.2 が公開されていない可能性）", file=sys.stderr)
        return None

    total_col = None
    for col in range(5, ws.max_column + 1):
        v = ws.cell(row=6, column=col).value
        if v and '合計' in str(v):
            total_col = col
            break

    results = []
    parent_state = [None]

    for row in range(8, ws.max_row + 1):
        raw_name = ws.cell(row=row, column=2).value
        area = extract_house_area(raw_name, parent_state)
        if not area:
            continue

        team_votes = parse_number(ws.cell(row=row, column=team_col).value)
        if team_votes is None:
            continue
        team_votes = round(team_votes)

        total = parse_number(ws.cell(row=row, column=total_col).value) if total_col else None
        rate = round(team_votes / total * 100, 1) if total and total > 0 else 0

        results.append({
            '地域': area,
            '得票数': team_votes,
            '得票率': rate,
            '合計': round(total) if total else 0,
        })

    return results


# ============================================================
# 衆院比例（P_20号様式シート）パーサー — 福岡等の選管Excel形式
# ============================================================

def parse_house_p20(filepath):
    """
    衆院比例の P_20号様式 シートをパースする。
    「開票調」シートがないExcel（福岡県等）向けフォールバック。

    想定フォーマット:
    - Sheet: 'P_20号様式' (名前に 'P_20' を含む)
    - Row 1: ヘッダー (頁番号, 行番号, 市区町村名, [届出番号N, 政党名N, 得票数N]..., 小計, ...)
    - Row 2+: データ行
    - 奇数ページ: 主要政党(10党), 偶数ページ: 追加政党(日本共産党等)
    - 政令指定都市の区は直接列挙され、＊{市名}計 で小計行がある
    - 選挙区分割 (例: 東区（１区）, 東区（４区）) は合算する
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = None
    for sn in wb.sheetnames:
        if 'P_20' in sn:
            ws = wb[sn]
            break
    if ws is None:
        return None

    # チームみらいの得票数列を探す
    team_vote_col = None
    subtotal_col = None
    for col in range(4, ws.max_column + 1):
        h = str(ws.cell(1, col).value or '')
        if h.startswith('政党名'):
            pname = str(ws.cell(2, col).value or '')
            if 'チームみらい' in pname:
                team_vote_col = col + 1  # 次の列が得票数
        elif h == '小計':
            subtotal_col = col
            break

    if team_vote_col is None:
        print(f"  Warning: P_20号様式にチームみらい列が見つかりません", file=sys.stderr)
        return None

    # 全データ行を収集（奇数ページ: 主要政党, 偶数ページ: 追加政党）
    odd_rows = []   # (raw_name, team_votes, subtotal)
    even_rows = []  # (raw_name, extra_votes)

    for r in range(2, ws.max_row + 1):
        raw_name = str(ws.cell(r, 3).value or '').strip()
        if not raw_name or raw_name == 'None':
            continue
        page = int(ws.cell(r, 1).value or 0)
        if page % 2 == 1:
            team = parse_number(ws.cell(r, team_vote_col).value)
            sub = parse_number(ws.cell(r, subtotal_col).value) if subtotal_col else 0
            odd_rows.append((raw_name, team, sub or 0))
        else:
            # 偶数ページ: 追加政党の得票数を全て合算
            extra_total = 0
            for ec in range(6, 34, 3):  # 得票数1, 得票数2, ... (cols 6, 9, 12, ...)
                v = parse_number(ws.cell(r, ec).value)
                if v is not None:
                    extra_total += v
            even_rows.append((raw_name, extra_total if extra_total > 0 else None))

    def _clean(name):
        return name.replace('＊', '').replace('*', '').replace(' ', '').replace('\u3000', '').strip()

    def _build_city_map(rows):
        """＊{市名}計 行を走査し、直前の区をその市に紐付ける city_map を返す。"""
        city_map = {}
        for i in range(len(rows)):
            clean = _clean(rows[i][0])
            if clean.endswith('計'):
                city = clean.replace('計', '')
                if city in DESIGNATED_CITIES:
                    for j in range(i - 1, -1, -1):
                        prev_clean = _clean(rows[j][0])
                        if prev_clean.endswith('計'):
                            break
                        if j in city_map:
                            break
                        city_map[j] = city
        return city_map

    city_map_odd = _build_city_map(odd_rows)
    city_map_even = _build_city_map(even_rows)

    def _to_area(raw_name, city):
        """生のエリア名を正規化。分割区(東区（１区）)は区名のみにする。"""
        clean = _clean(raw_name)
        if city:
            ward = re.sub(r'[（(].+?[）)]', '', clean)
            return city + ward
        return clean

    # 奇数ページ → merged dict
    merged = {}
    for i, (raw_name, team, sub) in enumerate(odd_rows):
        clean = _clean(raw_name)
        if clean.endswith('計') or clean.endswith('郡'):
            continue
        if team is None:
            continue
        area = _to_area(raw_name, city_map_odd.get(i))
        if area in merged:
            merged[area]['team'] += round(team)
            merged[area]['subtotal'] += sub
        else:
            merged[area] = {'team': round(team), 'subtotal': sub, 'extra': 0}

    # 偶数ページ → extra votes を加算
    for i, (raw_name, extra) in enumerate(even_rows):
        clean = _clean(raw_name)
        if clean.endswith('計') or clean.endswith('郡'):
            continue
        if extra is None:
            continue
        area = _to_area(raw_name, city_map_even.get(i))
        if area in merged:
            merged[area]['extra'] += round(extra)

    results = []
    for area, d in merged.items():
        total = d['subtotal'] + d['extra']
        rate = round(d['team'] / total * 100, 1) if total > 0 else 0
        results.append({
            '地域': area,
            '得票数': d['team'],
            '得票率': rate,
            '合計': round(total),
        })

    return results


# ============================================================
# 衆院比例（PDF版）パーサー — r8hirei.pdf形式
# ============================================================

def _normalize_pdf_area(name):
    """PDFセル中の自治体名を正規化（半角・全角空白を除去）"""
    if not name:
        return ''
    return str(name).replace(' ', '').replace('\u3000', '').replace('\xa0', '').strip()


def parse_house_pdf(filepath):
    """
    衆院比例（党派別開票区別）のPDFをパースする。
    神奈川県選管 r8hirei.pdf 形式に対応。

    PDFには2種類のテーブルが含まれる:
    - 「党派別開票区別得票数」テーブル（複数ページに政党を分割）
    - 「開票区別投票総数」テーブル（合計列を取得するため）
    """
    if not HAS_PDFPLUMBER:
        print("  Error: pdfplumber が必要です。 pip install pdfplumber でインストールしてください。", file=sys.stderr)
        return None

    municipality_totals = {}  # area → total votes
    party_data = {}           # area → {party_name: votes}

    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 3:
                    continue

                header = table[0]
                # 投票総数テーブル判定: 「得票総数」を含む列がある
                is_total_table = any(c and '得票総数' in str(c) for c in header)
                # 党派別テーブル判定: 2行目に党名がある
                second_row = table[1] if len(table) > 1 else []
                is_party_table = any(
                    c and any(kw in str(c) for kw in ['みらい', '民主党', '共産党', '維新', '保守', 'れいわ', '参政', '改革', '社会民主', '公明'])
                    for c in second_row
                )

                if is_total_table:
                    # 開票区名 → 得票総数（Ａ）の列を特定
                    area_col = 0
                    total_col = None
                    for i, c in enumerate(header):
                        if c and '得票総数' in str(c):
                            total_col = i
                            break
                    if total_col is None:
                        continue
                    for row in table[1:]:
                        if len(row) <= max(area_col, total_col):
                            continue
                        area = _normalize_pdf_area(row[area_col])
                        total_str = (row[total_col] or '').replace(',', '').strip()
                        if not area or not total_str:
                            continue
                        try:
                            municipality_totals[area] = int(total_str)
                        except ValueError:
                            continue

                elif is_party_table:
                    # 党名行から各列の党名を取得
                    party_names = [str(c).strip() if c else '' for c in second_row]
                    # データ行
                    for row in table[2:]:
                        if not row or not row[0]:
                            continue
                        area = _normalize_pdf_area(row[0])
                        if not area or area in ('開票区名',):
                            continue
                        if area not in party_data:
                            party_data[area] = {}
                        for col_idx in range(1, len(row)):
                            if col_idx >= len(party_names):
                                break
                            pname = party_names[col_idx]
                            if not pname:
                                continue
                            val_str = (row[col_idx] or '').replace(',', '').strip()
                            if not val_str:
                                continue
                            try:
                                party_data[area][pname] = int(val_str)
                            except ValueError:
                                continue

    # チームみらいのデータを抽出
    results = []
    for area, parties in party_data.items():
        if 'チームみらい' not in parties:
            continue

        # 政令指定都市の親市行はスキップ（名前単独で「横浜市」「川崎市」など）
        if area in DESIGNATED_CITIES:
            continue

        # 郡名のみの行はスキップ
        if is_gun(area):
            continue

        # 集計行はスキップ（県計、指定都市計、その他市計、町村計）
        if is_total_row(area):
            continue

        # 郡名前置を除去
        clean_area = strip_gun_prefix(area)

        team_votes = parties['チームみらい']
        total = municipality_totals.get(area)
        rate = round(team_votes / total * 100, 1) if total and total > 0 else 0

        results.append({
            '地域': clean_area,
            '得票数': team_votes,
            '得票率': rate,
            '合計': total or 0,
        })

    return results


# ============================================================
# CSV出力 + ファイル探索
# ============================================================

def write_csv(rows, filepath, fieldnames):
    os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, '') for k in fieldnames})
    print(f"  → {filepath} ({len(rows)} rows)", file=sys.stderr)


def find_files(raw_dir, exts, *required_keywords):
    """raw_dir 内で指定拡張子かつ全キーワードを含むファイルを探す"""
    matches = []
    for ext in exts:
        for path in sorted(glob.glob(os.path.join(raw_dir, f'*.{ext}'))):
            name = os.path.basename(path).lower()
            if all(kw.lower() in name for kw in required_keywords):
                matches.append(path)
    return matches


def find_excel_files(raw_dir, *required_keywords):
    return find_files(raw_dir, ['xlsx', 'xls'], *required_keywords)


def find_pdf_files(raw_dir, *required_keywords):
    return find_files(raw_dir, ['pdf'], *required_keywords)


def convert_prefecture(prefecture_dir):
    raw_dir = os.path.join(prefecture_dir, 'raw')
    data_dir = os.path.join(prefecture_dir, 'data')

    if not os.path.isdir(raw_dir):
        print(f"Error: raw/ ディレクトリが見つかりません: {raw_dir}", file=sys.stderr)
        sys.exit(1)

    print(f"Processing: {prefecture_dir}", file=sys.stderr)
    print(f"  raw/  → {raw_dir}", file=sys.stderr)
    print(f"  data/ → {data_dir}", file=sys.stderr)
    print()

    converted = 0

    # 参院選挙区（候補者別） → candidate.csv
    files = find_excel_files(raw_dir, 'sangi', 'senkyoku')
    for f in files:
        print(f"参院選挙区: {os.path.basename(f)}", file=sys.stderr)
        results = parse_senate_district(f)
        if results:
            write_csv(results, os.path.join(data_dir, 'candidate.csv'), ['地域', '得票', '率'])
            converted += 1
        else:
            print(f"  Warning: パース失敗（チームみらい列が見つかりません）", file=sys.stderr)

    # 参院比例（党派別開票区別） → senate.csv
    files = find_excel_files(raw_dir, 'sangi', 'hirei', 'district')
    for f in files:
        print(f"参院比例（党派別開票区別）: {os.path.basename(f)}", file=sys.stderr)
        results = parse_house_p20(f)  # P_20号様式 を優先（福岡等）
        if not results:
            results = parse_senate_proportional(f)  # 開票区別シートにフォールバック
        if results:
            write_csv(results, os.path.join(data_dir, 'senate.csv'), ['地域', '得票数', '得票率'])
            converted += 1
        else:
            print(f"  Warning: パース失敗", file=sys.stderr)

    # 衆院比例（Excel: 開票調シート → P_20号様式 フォールバック） → house.csv
    files = find_excel_files(raw_dir, 'shugi', 'hirei')
    house_done = False
    for f in files:
        print(f"衆院比例 (Excel): {os.path.basename(f)}", file=sys.stderr)
        results = parse_house_proportional(f)
        if not results:
            results = parse_house_p20(f)
        if results:
            write_csv(results, os.path.join(data_dir, 'house.csv'),
                      ['地域', '得票数', '得票率', '合計'])
            converted += 1
            house_done = True

    # 衆院比例（PDF版） → house.csv （Excelで取得済みでなければ）
    if not house_done:
        pdfs = find_pdf_files(raw_dir, 'shugi', 'hirei')
        for f in pdfs:
            print(f"衆院比例 (PDF): {os.path.basename(f)}", file=sys.stderr)
            results = parse_house_pdf(f)
            if results:
                write_csv(results, os.path.join(data_dir, 'house.csv'),
                          ['地域', '得票数', '得票率', '合計'])
                converted += 1

    if converted == 0:
        print("\nWarning: 変換可能なExcelファイルが見つかりませんでした。", file=sys.stderr)
        sys.exit(1)

    print(f"\nDone! {converted} file(s) converted.", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(
        description='選挙管理委員会 Excel/PDF → CSV 変換ツール (Fork方式)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用例:
  python3 tools/convert_excel.py

ディレクトリ構造（リポジトリルート直下）:
  ./
  ├── raw/                                  ← 選管Excel/PDFをここに配置
  │   ├── R7.7_sangi_senkyoku.xlsx          → candidate.csv
  │   ├── R7.7_sangi_hirei_district.xlsx    → senate.csv
  │   └── R8.2_shugi_hirei.pdf              → house.csv
  └── data/                                 ← CSV自動生成
      ├── candidate.csv
      ├── senate.csv
      └── house.csv

ファイル命名規則（ファイル名にキーワードを含める・小文字化判定）:
  - sangi + senkyoku           → 参院選挙区候補者別      → candidate.csv
  - sangi + hirei + district   → 参院比例党派別開票区別  → senate.csv
  - shugi + hirei (.xlsx/.pdf) → 衆院比例党派別開票区別  → house.csv
        """,
    )
    parser.add_argument(
        '--root',
        default='.',
        help='リポジトリルート (default: 現在のディレクトリ)',
    )

    args = parser.parse_args()
    convert_prefecture(args.root)


if __name__ == '__main__':
    main()
